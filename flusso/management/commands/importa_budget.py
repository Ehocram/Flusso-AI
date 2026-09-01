"""Importa i workbook Budget ed Extra Budget conservando TUTTI i fogli e le colonne.

Uso (dal server, con i file caricati su disco):
    python manage.py importa_budget --budget /percorso/Budget_2026_V5.xlsx \\
                                    --extra /percorso/Extra_Budget_2026.xlsx --anno 2026

Ogni foglio del file diventa un FoglioBudget con le sue intestazioni originali e le
sue righe (RigaBudget). Il foglio «Budget» e il foglio dell'extra budget sono marcati
come principali: sono quelli su cui vengono copiati i progetti approvati.

L'import e' idempotente per (chiave, anno): rilanciandolo le righe importate dal file
vengono sostituite. Le righe generate dai progetti (da_progetto=True) sono PRESERVATE:
non si perdono ricaricando il file.
"""

import re
import unicodedata

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from flusso.models import FoglioBudget, RigaBudget, TipoFoglio


def _slug(testo: str) -> str:
    testo = unicodedata.normalize("NFKD", str(testo)).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", testo.lower()).strip("-")[:60] or "foglio"


def _valore(v):
    """Valore serializzabile in JSON, con le date in formato leggibile."""
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    return v


class Command(BaseCommand):
    help = "Importa i fogli dei workbook Budget ed Extra Budget nell'applicazione."

    def add_arguments(self, parser):
        parser.add_argument("--budget", help="Percorso del file Budget (.xlsx)")
        parser.add_argument("--extra", help="Percorso del file Extra Budget (.xlsx)")
        parser.add_argument("--anno", type=int, default=2026)

    def handle(self, *args, **opts):
        try:
            import openpyxl  # noqa
        except ImportError:
            raise CommandError("openpyxl non installato: pip install openpyxl")
        if not opts["budget"] and not opts["extra"]:
            raise CommandError("Indica almeno --budget o --extra.")
        tot_fogli = tot_righe = 0
        if opts["budget"]:
            f, r = self._importa(opts["budget"], opts["anno"], principale="Budget",
                                 tipo_principale=TipoFoglio.BUDGET, prefisso="")
            tot_fogli += f
            tot_righe += r
        if opts["extra"]:
            f, r = self._importa(opts["extra"], opts["anno"], principale=None,
                                 tipo_principale=TipoFoglio.EXTRA, prefisso="xb-")
            tot_fogli += f
            tot_righe += r
        self.stdout.write(self.style.SUCCESS(
            f"Import completato: {tot_fogli} fogli, {tot_righe} righe."))

    # ------------------------------------------------------------------
    def _importa(self, percorso, anno, principale, tipo_principale, prefisso):
        import openpyxl

        try:
            wb = openpyxl.load_workbook(percorso, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File non trovato: {percorso}")
        n_fogli = n_righe = 0
        for indice, nome in enumerate(wb.sheetnames):
            ws = wb[nome]
            righe = [[_valore(v) for v in r] for r in ws.iter_rows(values_only=True)]
            # Taglia le righe e le colonne completamente vuote in coda.
            while righe and not any(str(v).strip() for v in righe[-1]):
                righe.pop()
            if not righe:
                self.stdout.write(self.style.WARNING(f"  [vuoto] {nome}: saltato"))
                continue
            i_hdr = self._riga_intestazione(righe)
            intest = [str(v).strip() for v in righe[i_hdr]]
            while intest and not intest[-1]:
                intest.pop()
            corpo_grezzo = righe[i_hdr + 1:]
            larghezza_corpo = 0
            for r in corpo_grezzo:
                ultimo = 0
                for j, v in enumerate(r):
                    if str(v).strip():
                        ultimo = j + 1
                larghezza_corpo = max(larghezza_corpo, ultimo)
            n_col = max(len(intest), larghezza_corpo, 1)
            intest = intest + [f"Col {i + 1}" for i in range(len(intest), n_col)]
            corpo = [list(r[:n_col]) + [""] * max(0, n_col - len(r[:n_col]))
                     for r in corpo_grezzo if any(str(v).strip() for v in r[:n_col])]

            unico = len(wb.sheetnames) == 1
            tipo = (tipo_principale if (nome == principale or (unico and principale is None))
                    else TipoFoglio.SUPPORTO)
            chiave = f"{prefisso}{_slug(nome)}-{anno}"
            with transaction.atomic():
                foglio, _ = FoglioBudget.objects.update_or_create(
                    chiave=chiave,
                    defaults={"nome": nome, "tipo": tipo, "anno": anno,
                              "intestazioni": intest, "ordine": indice},
                )
                # Le righe generate dai progetti non si toccano.
                foglio.righe.filter(da_progetto=False).delete()
                base = (foglio.righe.count() + 1) * 1000
                RigaBudget.objects.bulk_create([
                    RigaBudget(foglio=foglio, ordine=base + i, dati=list(r))
                    for i, r in enumerate(corpo)
                ])
            n_fogli += 1
            n_righe += len(corpo)
            etichetta = "PRINCIPALE" if tipo != TipoFoglio.SUPPORTO else "supporto"
            self.stdout.write(f"  [{etichetta}] {nome}: {len(corpo)} righe, {n_col} colonne")
        wb.close()
        return n_fogli, n_righe

    @staticmethod
    def _riga_intestazione(righe) -> int:
        """Indice della riga di intestazione: fra le prime, quella con piu' celle piene.

        I workbook reali hanno righe di titolo/gruppo sopra l'intestazione vera
        (es. «Budget», «Scheduling», totali): sono poche celle sparse, mentre la
        riga di intestazione e' la piu' densa. A parita' vince la prima.
        """
        migliore, punteggio = 0, -1
        for i, r in enumerate(righe[:12]):
            piene = sum(1 for v in r if str(v).strip())
            testuali = sum(1 for v in r if isinstance(v, str) and v.strip())
            if piene > punteggio and testuali >= 3:
                migliore, punteggio = i, piene
        return migliore
