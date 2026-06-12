"""
Import dell'elenco progetti AI da Excel.

Legge il file `Elenco_Progetti_AI.xlsx` e crea le richieste con i dati della scheda
(titolo, area, descrizione, soluzione proposta, costo, ritorni, referente).

Tutti i progetti vengono creati nello stato **"Inviata alla Funzione AI"**: nel nuovo
sistema il flusso (qualifica, approvazione, avvio, monitoraggio) viene rieseguito da capo,
a prescindere dalle approvazioni informali pregresse riportate nel file. L'invio iniziale
è registrato nell'audit trail con l'owner di funzione come attore.
"""

from pathlib import Path

from accounts.models import Funzione
from flusso.models import Richiesta
from flusso.workflow import Stato, puo_eseguire

# Percorso del file Excel incluso nel progetto (default).
DEFAULT_XLSX = Path(__file__).resolve().parent / "data" / "Elenco_Progetti_AI.xlsx"

# Mappa "Richiedente" (colonna Excel) -> codice Funzione applicativo.
FUNZIONE_DA_RICHIEDENTE = {
    "IT": Funzione.IT,
    "R&D": Funzione.RND,
    "SALES": Funzione.SALES,
    "OPERATIONS": Funzione.OPERATIONS,
    "SUPPLY CHAIN": Funzione.SUPPLY_CHAIN,
    "HR": Funzione.HR,
    "FINANCE": Funzione.FINANCE,
}

# Indici di colonna (0-based) nelle righe dati.
COL = {
    "id": 0, "richiedente": 1, "titolo": 3, "descrizione": 4, "soluzione": 11,
    "ritorno_eco": 12, "ritorno_qual": 13, "referente": 14, "investimento": 16,
    "data_pres": 17, "approvazione": 18, "sal": 21,
}

# Percorsi di workflow (azione:attore) per raggiungere ogni stato target.
_BASE = ["invia:owner", "prendi_in_carico:ai", "presenta_approvazione:ai"]
PERCORSI = {
    Stato.INVIATA: ["invia:owner"],
    Stato.IN_APPROVAZIONE: _BASE,
    Stato.ATTIVO: _BASE + ["approva:appr", "avvia_progetto:ai"],
    Stato.MONITORAGGIO: _BASE + ["approva:appr", "avvia_progetto:ai", "avvia_monitoraggio:ai"],
}


def _norm(v) -> str:
    """Normalizza una cella in stringa pulita (spazi/newline compattati)."""
    if v is None:
        return ""
    return " ".join(str(v).split()).strip()


def _tronca(v, n: int) -> str:
    s = _norm(v)
    if len(s) <= n:
        return s
    taglio = s[:n].rsplit(" ", 1)[0]
    return (taglio or s[:n]).rstrip() + "…"


def _split_soluzione(sol: str):
    """'Frase soluzione — Tipo' -> (frase, tipo). Senza separatore: (frase, '')."""
    s = _norm(sol)
    if not s:
        return "", ""
    for sep in (" — ", "—", " – ", "–", " - "):
        if sep in s:
            a, b = s.split(sep, 1)
            return a.strip(), b.strip()
    return s, ""


def _percentuale(sal) -> int:
    """SAL in frazione (0..1) o percentuale (0..100) -> intero 0..100."""
    if sal in (None, ""):
        return 0
    try:
        v = float(sal)
    except (TypeError, ValueError):
        return 0
    if v <= 1:
        v *= 100
    return max(0, min(100, round(v)))


def _stato_e_sal(approvazione: str, sal, soluzione: str):
    # Tutti i progetti partono come "Inviata alla Funzione AI": nel nuovo sistema
    # qualifica, approvazione, avvio e SAL vengono (ri)eseguiti da capo dal flusso.
    return Stato.INVIATA, 0


def leggi_excel(path=DEFAULT_XLSX) -> list[dict]:
    """Estrae e normalizza le righe dell'elenco progetti dal file Excel."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    def righe(foglio):
        ws = wb[foglio]
        out = {}
        for r in ws.iter_rows(values_only=True):
            if not r or r[0] is None:
                continue
            try:
                idv = int(r[0])
            except (TypeError, ValueError):
                continue
            out[idv] = r
        return out

    nomi = wb.sheetnames
    foglio_base = nomi[0]
    base = righe(foglio_base)
    # Eventuale foglio "attivati": sovrascrive approvazione/SAL dove valorizzati.
    foglio_attivati = next((n for n in nomi if "attiv" in n.lower()), None)
    attivati = righe(foglio_attivati) if foglio_attivati else {}

    progetti = []
    for idv in sorted(base):
        r = base[idv]
        a = attivati.get(idv)

        def c(chiave, riga=r):
            i = COL[chiave]
            return riga[i] if riga and i < len(riga) else None

        richiedente = _norm(c("richiedente")).upper()
        funzione = FUNZIONE_DA_RICHIEDENTE.get(richiedente)
        if funzione is None:
            continue  # riga non riconducibile a una funzione nota

        # Approvazione/SAL: priorità al foglio "attivati" se presenti.
        approvazione = c("approvazione")
        sal = c("sal")
        if a is not None:
            if _norm(c("approvazione", a)):
                approvazione = c("approvazione", a)
            if c("sal", a) not in (None, ""):
                sal = c("sal", a)

        soluzione = c("soluzione")
        frase, tipo = _split_soluzione(soluzione)
        stato, perc = _stato_e_sal(approvazione, sal, soluzione)

        descrizione = frase if frase else _tronca(c("descrizione"), 300)

        progetti.append({
            "numero": idv,
            "funzione": funzione,
            "titolo": _tronca(c("titolo"), 138) or f"Progetto {idv}",
            "tipo_soluzione": _tronca(tipo, 138),
            "descrizione": descrizione or "—",
            "costo": _tronca(c("investimento"), 78) or "TBD",
            "saving_economico": _tronca(c("ritorno_eco"), 78) or "TBD",
            "saving_qualitativo": _tronca(c("ritorno_qual"), 118),
            "referente_area": _tronca(c("referente"), 118),
            "stato": stato,
            "sal": perc,
        })
    return progetti


def importa(attori_owner: dict, chiavi: dict, path=DEFAULT_XLSX, *, log=lambda m: None) -> int:
    """
    Crea le richieste a partire dall'Excel.

    attori_owner: {codice_funzione: utente_owner}
    chiavi:       {'ai': utente, 'appr': utente}
    Ritorna il numero di richieste create (salta quelle gia' presenti).
    """
    creati = 0
    for p in leggi_excel(path):
        if Richiesta.objects.filter(numero=p["numero"]).exists():
            continue
        owner = attori_owner.get(p["funzione"])
        if owner is None:
            continue
        r = Richiesta(
            numero=p["numero"], funzione=p["funzione"], titolo=p["titolo"],
            tipo_soluzione=p["tipo_soluzione"], descrizione=p["descrizione"],
            costo=p["costo"], saving_economico=p["saving_economico"],
            saving_qualitativo=p["saving_qualitativo"], referente_area=p["referente_area"],
            proponente=owner,
        )
        r.save()
        _percorri(r, PERCORSI[p["stato"]], owner, chiavi)
        if p["stato"] == Stato.MONITORAGGIO and p["sal"]:
            r.aggiorna_sal(p["sal"], attore=chiavi["ai"], nota="Avanzamento da SAL periodico")
        creati += 1
        log(f"  · {r.codice} [{r.get_funzione_display()}] {r.titolo} → {r.get_stato_display()}"
            + (f" ({p['sal']}%)" if p["sal"] else ""))
    return creati


def _percorri(richiesta, percorso, owner, chiavi):
    mappa = {"owner": owner, **chiavi}
    for passo in percorso:
        azione, chiave = passo.split(":")
        attore = mappa[chiave]
        if not puo_eseguire(richiesta, attore, azione):
            raise RuntimeError(f"Percorso non valido su {richiesta.codice}: {azione} da {richiesta.stato}")
        richiesta.applica(azione, attore=attore)
